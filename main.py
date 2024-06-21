import requests
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date 
import os


def create_xlsv(data, filename):
    # Creating DataFrame
    df = pd.DataFrame(data[1:], columns=data[0])

    # Saving to Excel
    file_path = filename
    df.to_excel(file_path, index=False)

    # Adjusting the column width
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        worksheet = writer.sheets['Sheet1']
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            print
            col_letter = col[0].column_letter
            if col_letter == "B":
                worksheet.column_dimensions["B"].width = 20
            else:
                worksheet.column_dimensions[col_letter].width = max_length + 2


def convert_to_tb(filename):
    # Load the workbook and select the active sheet
    file_path = filename
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Define the table range and create a table
    table_range = f"A1:D{sheet.max_row}"
    table = Table(displayName="VulnerabilityTable", ref=table_range)

    # Add a table style
    style = TableStyleInfo(
        name="TableStyleMedium6",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style

    # Add the table to the sheet
    sheet.add_table(table)

    # Save the workbook
    workbook.save(file_path)

def unique_CVEs(filename, newfilename):
    # Step 4: Extract unique CVE titles and save to a new spreadsheet
    workbook = load_workbook(filename)
    sheet = workbook.active
    cve_titles = set()
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
        cve_titles.add(row[0])

    cve_titles_list = list(cve_titles)
    df_cve_titles = pd.DataFrame(cve_titles_list, columns=['CVE Title'])
    new_file_path = newfilename
    df_cve_titles.to_excel(new_file_path, index=False)


def directory_checker(year, month):
    # Define the directory and folder names
    base_directory = 'spreadsheets'
    year_folder = year
    month_folder = month
    year_folder_path = os.path.join(base_directory, year_folder)
    month_folder_path = os.path.join(year_folder_path, month_folder)

    # Check if the year folder exists
    if not os.path.exists(year_folder_path):
        # If the year folder does not exist, create it
        os.makedirs(year_folder_path)
        print(f"Folder '{year_folder}' created in '{base_directory}' directory.")

    # Check if the month folder exists within the year folder
    if not os.path.exists(month_folder_path):
        # If the month folder does not exist, create it
        os.makedirs(month_folder_path)
        print(f"Folder '{month_folder}' created in '{year_folder}' directory.")
    else:
        print(f"Folder '{month_folder}' already exists in '{year_folder}' directory.")


def create_dataset(URL):

    #URL = "https://www.bleepingcomputer.com/news/microsoft/microsoft-february-2024-patch-tuesday-fixes-2-zero-days-73-flaws/"

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36'
    }

    response = requests.get(URL, headers=headers)
    html = BeautifulSoup(response.text, 'html.parser')

    #raw datas
    total = html.find_all('p')
    zero_day = html.find_all('h2')
    table_rows = html.find_all('tr')


    #save data per row
    td_list = []
    descriptions = {}

    for t in total:
        if "The total count of" in str(t):
            print(str(t).split()[4])
            descriptions['flaws']=int(str(t).split()[4])
            break

    #get the number of zero days
    if "zero-days" in str(zero_day[0]):
        print("There is zero day")
        soup_zero = BeautifulSoup(str(zero_day[0]), 'html.parser')
        zero_number = soup_zero.find('h2').get_text()
        descriptions['zero-days']=zero_number.split()[0]
    else: 
        print("No zero days")
        descriptions['zero-days']="Zero"


    for tr in table_rows:
        soup_td = BeautifulSoup(str(tr), 'html.parser')
        if "<th>" in str(tr):
            th_contents = [th.get_text() for th in soup_td.find_all('th')]
            td_list.append(th_contents)
        else:
            td_contents = [td.get_text() for td in soup_td.find_all('td')]
            if "CVE-" in str(td_contents):
                hyperlink = f'=HYPERLINK("https://msrc.microsoft.com/update-guide/en-US/advisory/{str(td_contents[1])}", "{str(td_contents[1])}")'
                td_contents[1]=hyperlink

            td_list.append(td_contents)


    create_xlsv(td_list, f'./spreadsheets/{current_year}/{current_month}/Patch-Tuesday-{current_month}-{current_year}.xlsx')
    convert_to_tb(f'./spreadsheets/{current_year}/{current_month}/Patch-Tuesday-{current_month}-{current_year}.xlsx')
    unique_CVEs(f'./spreadsheets/{current_year}/{current_month}/Patch-Tuesday-{current_month}-{current_year}.xlsx', f'./spreadsheets/{current_year}/{current_month}/CVE-Unique-{current_month}-{current_year}.xlsx')

    #count the occurences
    # Read the Excel file
    df = pd.read_excel(f'./spreadsheets/{current_year}/{current_month}/Patch-Tuesday-{current_month}-{current_year}.xlsx')

    # Count the number of occurrences of "Important" in the fourth column
    important_count = df.iloc[:, 3].value_counts().get('Important', 0)
    critical_count = df.iloc[:, 3].value_counts().get('Critical', 0)
    descriptions['important']=int(important_count)
    descriptions['critical']=int(critical_count)
    return descriptions

def write_an_email(headline, dictionary):
    df = pd.read_excel(f'./spreadsheets/{current_year}/{current_month}/CVE-Unique-{current_month}-{current_year}.xlsx')
    # Convert the DataFrame to a Markdown table
    markdown_table = df.to_markdown(index=False)

    latest_email = f"""
    # Latest Patch Notification Email to be Sent!

    ## {headline}

    Microsoft have released their monthly Patch Tuesday cycle for {current_month}. This has been assessed by the AUCloud Security Operations Team.
    Of note, there were **{dictionary['flaws']}** Vulnerabilities patched this month, with **({dictionary['critical']})** of these assessed as **CRITICAL**, **{dictionary['important']}**. There were **{dictionary['zero-days']} Zero-Days** vulnerabilities.

    A list of affected products in **{current_month}** is as follows:
    {markdown_table}

    Please see the attched spreadsheet for further information regarding the vulnerabilities.

    AUCloud SOC have detection mechanism in place and with other Essential8 (E8) mitigations, do not believe this warrants a 48hr-patch window. All critical vulnerabilities must be patched within 2 weeks and the others can be patched within 4 weeks. AUCloud recommend patching completed of all Windows infrastructure within 2 Weeks.

    Please note, all recommendations are subject to change based on new information being disclosed on these patches.

    If you have any questions regarding this advisory, please contact AUCloud Security Operations.

    """

    md_file_path = f'./spreadsheets/{current_year}/{current_month}/Email-Notification.md'
    with open(md_file_path, 'w') as file:
        file.write(latest_email)


URL = "https://www.bleepingcomputer.com/tag/patch-tuesday/"

headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36'
    }

response = requests.get(URL, headers=headers)
html = BeautifulSoup(response.text, 'html.parser')

headlines = html.find_all('h4')

todays_date = date.today() 
current_month = todays_date.strftime("%B")
current_year = todays_date.strftime("%Y")

directory_checker(str(current_year), str(current_month))

for patch in headlines:
    if "Patch Tuesday" in str(patch):
        if str(current_month) in str(patch) and str(current_year) in str(patch):
            soup_headline = BeautifulSoup(str(patch), 'html.parser')
            h4_contents = [h4.get_text() for h4 in soup_headline.find_all('h4')]
            link = soup_headline.find('a')['href']


#check if file is already created
directory_path = os.path.join('spreadsheets', current_year, current_month)

if os.path.exists(directory_path):
    if any(os.scandir(directory_path)):
        print(f"The directory '{directory_path}' contains files.")
    else:
        print(f"The directory '{directory_path}' is empty.")
        print("creating files based on latest patch...")
        dataset_values = create_dataset(str(link))
        print(dataset_values)
        write_an_email(h4_contents, dataset_values)
else:
    print(f"The directory '{directory_path}' does not exist.")


